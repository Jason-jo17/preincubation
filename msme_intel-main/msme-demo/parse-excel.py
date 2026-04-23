import json
import os
import openpyxl

excel_path = r"D:\Downloads2\Maharashtra_Tech_Ecosystem_Complete (1).xlsx"
out_path = r"d:\Buisness intel\msme-intelligence-platform\msme-demo\lib\demo-data\maharashtra-ecosystem.ts"

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

headers = []
for cell in ws[1]:
    headers.append(cell.value.strip() if cell.value else "")

cat_idx = -1
for i, h in enumerate(headers):
    if "Category" in h:
        cat_idx = i
        break

def get_col(row, name):
    for i, h in enumerate(headers):
        if name.lower() in h.lower() and i < len(row):
            return str(row[i].value).strip() if row[i].value else ""
    return ""

cats = {
    "Product Startup": {"id": "product-startup", "title": "Product Startups", "icon": "Rocket", "desc": "Hardware and software startups"},
    "AI/ML Company": {"id": "ai-ml", "title": "AI & ML Specialists", "icon": "Cpu", "desc": "Core AI/ML providers"},
    "AI/ML Services": {"id": "ai-ml", "title": "AI & ML Specialists", "icon": "Cpu", "desc": "Core AI/ML providers"},
    "Software Services": {"id": "software-services", "title": "Software Services", "icon": "Code", "desc": "Application development"},
    "Industrial Automation": {"id": "industry-40", "title": "Industry 4.0 & Automation", "icon": "Factory", "desc": "Robotics and smart manufacturing"},
    "Robotics": {"id": "industry-40", "title": "Industry 4.0 & Automation", "icon": "Factory", "desc": "Robotics and smart manufacturing"},
    "Electronics Design": {"id": "electronics", "title": "Electronics & Hard-Tech", "icon": "Cpu", "desc": "PCB and electronics"},
    "PCB Manufacturing": {"id": "electronics", "title": "Electronics & Hard-Tech", "icon": "Cpu", "desc": "PCB and electronics"},
    "Contract Manufacturing": {"id": "electronics", "title": "Electronics & Hard-Tech", "icon": "Cpu", "desc": "PCB and electronics"},
    "3D Printing": {"id": "manufacturing", "title": "Advanced Manufacturing", "icon": "Layers", "desc": "3D printing and prototyping"},
    "IoT Platform": {"id": "iot", "title": "IoT & Edge Computing", "icon": "Wifi", "desc": "IoT platforms and solutions"},
    "IoT Solutions": {"id": "iot", "title": "IoT & Edge Computing", "icon": "Wifi", "desc": "IoT platforms and solutions"},
    "Workflow Automation": {"id": "dt", "title": "Digital Transformation", "icon": "TrendingUp", "desc": "RPA and enterprise processes"},
    "RPA Services": {"id": "dt", "title": "Digital Transformation", "icon": "TrendingUp", "desc": "RPA and enterprise processes"},
    "ERP Implementation": {"id": "erp", "title": "Enterprise Resource Planning", "icon": "Database", "desc": "ERP implementation partners"},
    "Custom ERP": {"id": "erp", "title": "Enterprise Resource Planning", "icon": "Database", "desc": "ERP implementation partners"},
    "Tally Partner": {"id": "erp", "title": "Enterprise Resource Planning", "icon": "Database", "desc": "ERP implementation partners"},
    "Cloud/DevOps": {"id": "cloud", "title": "Cloud & Infrastructure", "icon": "Cloud", "desc": "DevOps and cloud migration"},
    "Digital Transformation": {"id": "dt", "title": "Digital Transformation", "icon": "TrendingUp", "desc": "Digital transformation experts"},
    "Incubator Portfolio": {"id": "incubators", "title": "Incubator Portfolios", "icon": "Lightbulb", "desc": "Leading deep-tech incubators"},
    "Tier-2 Flagship": {"id": "tier-2", "title": "Regional Leaders", "icon": "MapPin", "desc": "Regional tech companies"},
    "Tier-2 Company": {"id": "tier-2", "title": "Regional Leaders", "icon": "MapPin", "desc": "Regional tech companies"},
    "Notable Startup": {"id": "startups", "title": "Notable Startups", "icon": "Star", "desc": "Fast-growing companies"}
}

grouped = {}
for row_idx, row in enumerate(ws.iter_rows(min_row=2)):
    if not row[0].value:
        continue
    raw = get_col(row, 'Category')
    c = cats.get(raw, {"id": "other", "title": raw if raw else "Other", "icon": "Briefcase", "desc": "Other services"})
    cid = c["id"]
    if cid not in grouped:
        grouped[cid] = c.copy()
        grouped[cid]["companies"] = []
    
    services_str = get_col(row, 'Tech Stack')
    services = [s.strip().replace("'", "\\'") for s in services_str.split(',')] if services_str else []
    
    comp_name = get_col(row, 'Company Name').replace("'", "\\'")
    comp_desc = get_col(row, 'Product/Service').replace("'", "\\'")
    comp_loc = get_col(row, 'Location').replace("'", "\\'")
    comp_contact = get_col(row, 'Contact').replace("'", "\\'")
    comp_web = get_col(row, 'Website').replace("'", "\\'")
    if comp_web and not comp_web.startswith('http'):
        comp_web = f"https://{comp_web}"

    comp = {
        "id": f"{cid}-{len(grouped[cid]['companies'])+1}",
        "name": comp_name,
        "description": comp_desc,
        "location": comp_loc,
        "contact": {
            "email": comp_contact,
            "website": comp_web
        },
        "services": services
    }
    grouped[cid]["companies"].append(comp)

out = "import { EcosystemCategory } from './ecosystem-providers';\n\nexport const MAHARASHTRA_ECOSYSTEM: EcosystemCategory[] = [\n"
for g in grouped.values():
    out += f"    {{\n        id: '{g['id']}',\n        title: '{g['title']}',\n        description: '{g['desc']}',\n        icon_name: '{g['icon']}' as any,\n        companies: [\n"
    for c in g['companies']:
        out += f"            {{\n                id: '{c['id']}',\n                name: '{c['name']}',\n                description: '{c['description']}',\n                rating: 4.8,\n                location: '{c['location']}',\n                contact: {{\n                    email: '{c['contact']['email']}',\n                    phone: '',\n                    website: '{c['contact']['website']}'\n                }},\n                services: [{', '.join([f'{chr(39)}{s}{chr(39)}' for s in c['services'] if s])}],\n                match_score: 92\n            }},\n"
    out += "        ]\n    },\n"
out += "];\n"

with open(out_path, 'w', encoding='utf-8') as f:
    f.write(out)

print(f"Done processing {len(grouped.keys())} categories to {out_path}")
